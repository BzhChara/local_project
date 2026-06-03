from datetime import datetime

from openpyxl import load_workbook

from led_debugger.exporter import format_current_value, format_timestamp
from led_debugger.recorder import ContinuousRecorder, TEMP_CSV_DIR_NAME, default_recording_dir, prepare_recording_dir


def test_default_recording_dir_uses_start_time(tmp_path) -> None:
    moment = datetime(2026, 6, 1, 14, 30, 0)
    assert default_recording_dir(moment, tmp_path / "data") == tmp_path / "data" / "2026_06_01" / "record_143000"


def test_prepare_recording_dir_reuses_and_clears_same_second_dir(tmp_path) -> None:
    moment = datetime(2026, 6, 1, 14, 30, 0)
    output_dir = tmp_path / "data" / "2026_06_01" / "record_143000"
    stale_file = output_dir / "old.txt"
    stale_file.parent.mkdir(parents=True)
    stale_file.write_text("stale", encoding="utf-8")

    assert prepare_recording_dir(moment, tmp_path / "data") == output_dir
    assert not output_dir.exists()


def test_continuous_recorder_writes_interval_samples_and_exports_xlsx(tmp_path) -> None:
    output_dir = tmp_path / "data" / "2026_06_01" / "record_143000"
    recorder = ContinuousRecorder(output_dir, interval_seconds=1.0)
    values = (0.0950616002082825, *(float(index) for index in range(2, 17)))

    assert recorder.record_channel_reading(1, values, 100.0) == 16
    assert recorder.record_channel_reading(1, tuple(value + 1.0 for value in values), 100.5) == 0
    assert recorder.record_channel_reading(1, tuple(value + 2.0 for value in values), 101.2) == 16

    summary = recorder.finish()

    assert summary.output_dir == output_dir
    assert summary.files_written == 256
    assert summary.rows_written == 32
    assert (output_dir / TEMP_CSV_DIR_NAME / "通道1_LED电流1.csv").exists()

    workbook = load_workbook(output_dir / "通道1" / "LED电流1.xlsx")
    sheet = workbook.active
    assert sheet["A1"].value == "时间"
    assert sheet["B1"].value == "电流值(mA)"
    assert sheet["A2"].value == format_timestamp(100.0)
    assert sheet["B2"].value == format_current_value(0.0950616002082825)
    assert sheet["A3"].value == format_timestamp(101.2)
    assert sheet["B3"].value == format_current_value(0.0950616002082825 + 2.0)

    empty_workbook = load_workbook(output_dir / "通道16" / "LED电流16.xlsx")
    empty_sheet = empty_workbook.active
    assert empty_sheet["A1"].value == "时间"
    assert empty_sheet["A2"].value is None
