from datetime import datetime

from openpyxl import load_workbook

from led_debugger.data_store import DataStore, HistoryPoint
from led_debugger.exporter import (
    existing_export_files,
    export_all_histories,
    format_current_value,
    format_export_date,
    format_timestamp,
    required_history_points,
    select_points_by_interval,
)


def test_select_points_by_interval_uses_next_available_sample() -> None:
    points = [
        HistoryPoint(15.0, 1.0),
        HistoryPoint(15.9, 2.0),
        HistoryPoint(16.3, 3.0),
        HistoryPoint(17.0, 4.0),
        HistoryPoint(18.0, 5.0),
        HistoryPoint(18.8, 6.0),
    ]

    selected = select_points_by_interval(points, interval_seconds=1.0)

    assert [(point.timestamp, point.value_ma) for point in selected] == [
        (15.0, 1.0),
        (16.3, 3.0),
        (18.0, 5.0),
    ]


def test_required_history_points_rounds_up() -> None:
    assert required_history_points(duration_seconds=60.0, interval_seconds=1.0) == 61
    assert required_history_points(duration_seconds=10.0, interval_seconds=3.0) == 5


def test_format_current_value_keeps_full_text_representation() -> None:
    assert format_current_value(0.0950616002082825) == "0.0950616002082825"


def test_export_date_and_timestamp_formats() -> None:
    moment = datetime(2026, 5, 19, 14, 5, 25, 759000)

    assert format_export_date(moment) == "2026_05_19"
    assert format_timestamp(moment.timestamp()) == "14:05:25.759"


def test_export_all_histories_writes_channel_led_xlsx_files(tmp_path) -> None:
    store = DataStore(max_points=10)
    values = (0.0950616002082825, *(float(index) for index in range(2, 17)))
    store.add_channel_reading(channel=1, values_ma=values, timestamp=100.0)
    store.add_channel_reading(channel=1, values_ma=tuple(value + 10.0 for value in values), timestamp=101.3)
    output_dir = tmp_path / "data" / format_export_date(datetime(2026, 5, 19))

    summary = export_all_histories(
        store,
        output_dir,
        duration_seconds=60.0,
        interval_seconds=1.0,
    )

    led_1_xlsx = output_dir / "通道1" / "LED电流1.xlsx"
    assert summary.files_written == 256
    assert summary.rows_written == 32
    assert led_1_xlsx.exists()

    workbook = load_workbook(led_1_xlsx)
    sheet = workbook.active
    assert sheet.title == "LED电流"
    assert sheet.column_dimensions["A"].width == 24
    assert sheet.column_dimensions["B"].width == 22
    assert sheet["A1"].value == "时间"
    assert sheet["B1"].value == "电流值(mA)"
    assert sheet["A1"].font.bold is True
    assert sheet["B1"].font.bold is True
    assert sheet["A1"].font.name == "微软雅黑"
    assert sheet["B1"].font.name == "微软雅黑"
    assert sheet["A2"].value == format_timestamp(100.0)
    assert sheet["A2"].data_type == "s"
    assert sheet["A2"].number_format == "@"
    assert sheet["A2"].font.name == "微软雅黑"
    assert sheet["B2"].value == "0.0950616002082825"
    assert sheet["B2"].data_type == "s"
    assert sheet["B2"].number_format == "@"
    assert sheet["B2"].font.name == "微软雅黑"
    assert sheet["A3"].value == format_timestamp(101.3)
    assert sheet["B3"].value == format_current_value(0.0950616002082825 + 10.0)

    empty_workbook = load_workbook(output_dir / "通道16" / "LED电流16.xlsx")
    empty_sheet = empty_workbook.active
    assert empty_sheet["A1"].value == "时间"
    assert empty_sheet["A2"].value is None


def test_existing_export_files_only_checks_fixed_xlsx_names(tmp_path) -> None:
    output_dir = tmp_path / "data" / format_export_date(datetime(2026, 5, 19))
    led_xlsx = output_dir / "通道1" / "LED电流1.xlsx"
    unrelated = output_dir / "其他文件.xlsx"
    legacy_csv = output_dir / "通道1" / "LED电流1.csv"
    old_day_xlsx = tmp_path / "data" / "2026_05_18" / "通道1" / "LED电流1.xlsx"
    led_xlsx.parent.mkdir(parents=True)
    old_day_xlsx.parent.mkdir(parents=True)
    led_xlsx.write_text("old", encoding="utf-8")
    unrelated.write_text("keep", encoding="utf-8")
    legacy_csv.write_text("legacy", encoding="utf-8")
    old_day_xlsx.write_text("old day", encoding="utf-8")

    assert existing_export_files(output_dir) == [led_xlsx]
