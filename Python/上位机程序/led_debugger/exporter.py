from __future__ import annotations

import sys
from dataclasses import dataclass
from datetime import datetime
from math import ceil, isfinite
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from led_debugger.data_store import DataStore, HistoryPoint


EXPORT_CHANNELS = range(1, 17)
EXPORT_LEDS = range(1, 17)
EXCEL_FONT_NAME = "微软雅黑"


@dataclass(frozen=True)
class ExportSummary:
    """Excel 导出结果摘要。"""

    output_dir: Path
    files_written: int
    rows_written: int


def application_root() -> Path:
    """返回程序所在目录；打包后为 exe 同级目录，源码运行时为 app.py 所在目录。"""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(sys.argv[0]).resolve().parent


def default_data_dir() -> Path:
    """默认 data 输出目录。"""
    return application_root() / "data" / format_export_date()


def format_export_date(moment: datetime | None = None) -> str:
    """格式化导出目录日期。"""
    return (moment or datetime.now()).strftime("%Y_%m_%d")


def required_history_points(duration_seconds: float, interval_seconds: float) -> int:
    """按保存时长和间隔估算需要保留的最少历史点数。"""
    return ceil(duration_seconds / interval_seconds) + 1


def existing_export_files(output_dir: Path) -> list[Path]:
    """返回本程序固定导出的 Excel 文件中已经存在的部分。"""
    existing: list[Path] = []
    for channel in EXPORT_CHANNELS:
        for led_index in EXPORT_LEDS:
            path = _xlsx_path(output_dir, channel, led_index)
            if path.exists():
                existing.append(path)
    return existing


def has_exportable_history(data_store: DataStore) -> bool:
    """判断当前内存中是否存在任意可导出的历史数据。"""
    for channel in EXPORT_CHANNELS:
        for led_index in EXPORT_LEDS:
            if data_store.get_history(channel, led_index):
                return True
    return False


def export_all_histories(
    data_store: DataStore,
    output_dir: Path,
    duration_seconds: float,
    interval_seconds: float,
) -> ExportSummary:
    """按通道/LED 导出最近一段历史数据到 Excel。"""
    latest_timestamp = _latest_timestamp(data_store)
    if latest_timestamp is None:
        return ExportSummary(output_dir=output_dir, files_written=0, rows_written=0)

    output_dir.mkdir(parents=True, exist_ok=True)
    window_start = latest_timestamp - duration_seconds
    files_written = 0
    rows_written = 0

    for channel in EXPORT_CHANNELS:
        channel_dir = output_dir / f"通道{channel}"
        channel_dir.mkdir(parents=True, exist_ok=True)
        for led_index in EXPORT_LEDS:
            points = [
                point
                for point in data_store.get_history(channel, led_index)
                if window_start <= point.timestamp <= latest_timestamp
            ]
            selected_points = select_points_by_interval(points, interval_seconds)
            _write_xlsx(_xlsx_path(output_dir, channel, led_index), selected_points)
            files_written += 1
            rows_written += len(selected_points)

    return ExportSummary(output_dir=output_dir, files_written=files_written, rows_written=rows_written)


def select_points_by_interval(points: list[HistoryPoint], interval_seconds: float) -> list[HistoryPoint]:
    """从真实采样点中按间隔选择后续第一个满足时间的点。"""
    if not points:
        return []

    selected = [points[0]]
    target_timestamp = points[0].timestamp + interval_seconds
    index = 1
    while index < len(points):
        while index < len(points) and points[index].timestamp < target_timestamp:
            index += 1
        if index >= len(points):
            break
        selected.append(points[index])
        target_timestamp = points[index].timestamp + interval_seconds
        index += 1
    return selected


def format_timestamp(timestamp: float) -> str:
    """格式化真实采样时间，保留时分秒毫秒并按文本写入 Excel。"""
    return datetime.fromtimestamp(timestamp).strftime("%H:%M:%S.%f")[:-3]


def format_current_value(value_ma: float) -> str:
    """按文本导出电流值，避免 Excel 自动缩短小数显示。"""
    if not isfinite(value_ma):
        return f"{value_ma:g}"
    return str(value_ma)


def _latest_timestamp(data_store: DataStore) -> float | None:
    latest_timestamp: float | None = None
    for channel in EXPORT_CHANNELS:
        for led_index in EXPORT_LEDS:
            points = data_store.get_history(channel, led_index)
            if not points:
                continue
            point_timestamp = points[-1].timestamp
            if latest_timestamp is None or point_timestamp > latest_timestamp:
                latest_timestamp = point_timestamp
    return latest_timestamp


def _xlsx_path(output_dir: Path, channel: int, led_index: int) -> Path:
    return output_dir / f"通道{channel}" / f"LED电流{led_index}.xlsx"


def write_history_xlsx(path: Path, points: list[HistoryPoint]) -> None:
    """写入单个 LED 的历史数据 Excel。"""
    _write_xlsx(path, points)


def _write_xlsx(path: Path, points: list[HistoryPoint]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "LED电流"
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 22

    header_font = Font(name=EXCEL_FONT_NAME, bold=True)
    body_font = Font(name=EXCEL_FONT_NAME)
    sheet["A1"] = "时间"
    sheet["B1"] = "电流值(mA)"
    sheet["A1"].font = header_font
    sheet["B1"].font = header_font

    for row_index, point in enumerate(points, start=2):
        time_cell = sheet.cell(row=row_index, column=1, value=format_timestamp(point.timestamp))
        time_cell.number_format = "@"
        time_cell.font = body_font

        value_cell = sheet.cell(row=row_index, column=2, value=format_current_value(point.value_ma))
        value_cell.number_format = "@"
        value_cell.font = body_font

    workbook.save(path)
