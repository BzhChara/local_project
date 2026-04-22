import os
import io
import time
import ddddocr
import pandas as pd
from pywinauto import Desktop
from PIL import Image, ImageOps, ImageEnhance
from openpyxl.styles import Font

# 初始化
det = ddddocr.DdddOcr(beta=False, show_ad=False)

# 参数设置
Y_STARTS = [135, 278]
X_START = 98
WIDTH = 52
HEIGHT = 32
X_GAP = 54

DURATION = 360
SAMPLE_TIME = 5
INTERVAL = 0.5

def decimal(text):
    # 统一转小写
    text = text.lower()
    
    # 把所有可能被认错的字符强制转回数字
    # 常见误认：o/O -> 0, s -> 5, b -> 6, g -> 9, i/l -> 1, c -> 0
    mapping = {
        'o': '0',
        'c': '0',
        's': '5',
        'b': '6',
        'g': '9',
        'i': '1',
        'l': '1',
        'z': '2'
    }
    for char, num in mapping.items():
        text = text.replace(char, num)

    # 只提取数字
    nums = "".join([c for c in text if c.isdigit()])
    
    if not nums:
        return "0.000"
    
    # 保持识别到的数字，在倒数第3位前加点（例如 0123 -> 0.123）
    if len(nums) >= 4:
        return nums[:-3] + "." + nums[-3:]
    else:
        # 如果不满4位（比如漏读），前面补0对齐
        temp = nums.zfill(4)
        return temp[:-3] + "." + temp[-3:]

def ocr_process(img_roi, debug_path=None):
    # 四周切除，去除边框残留干扰
    w, h = img_roi.size
    img_roi = img_roi.crop((2, 2, w - 2, h - 2)) 
    
    # 放大并转灰度
    img_roi = img_roi.convert('L').resize((img_roi.width * 5, img_roi.height * 5), Image.LANCZOS)
    
    # 适度锐化
    img_roi = ImageEnhance.Sharpness(img_roi).enhance(5.0)
    
    # 二值化
    # 如果是点击开始检测记录屏幕会变灰，设置为115。直接记录设置为155
    img_roi = img_roi.point(lambda x: 0 if x < 155 else 255)

    # 反转颜色
    img_roi = ImageOps.invert(img_roi)
    
    img_byte_arr = io.BytesIO()
    img_roi.save(img_byte_arr, format='PNG')
    return det.classification(img_byte_arr.getvalue())

def run():
    duration = DURATION
    sample_time = SAMPLE_TIME
    interval = INTERVAL
    max_samples = int(sample_time / interval) # 10 个数据
    
    final_data = {f"CH{i+1:02d}": "0.000" for i in range(24)}
    # 存储每个通道的采样池
    sample_pools = {f"CH{i+1:02d}": [] for i in range(24)}
    
    print(f"开始采样...")
    start_time = time.time()

    try:
        file_name = "检测监控结果.xlsx"
        if os.path.exists(file_name):
            try:
                # 尝试重命名文件。如果文件被打开，Windows 不允许重命名，会抛出异常
                temp_name = "test_lock_temp.xlsx"
                os.rename(file_name, temp_name)
                os.rename(temp_name, file_name) # 成功了就换回来
            except OSError:
                print(f"\n错误。文件 '{file_name}' 正在被 Excel 或其他程序占用！")
                print("请先关闭该 Excel 文件，然后重新运行程序。")
                return
            
        app = Desktop(backend="win32")
        win = app.window(title_re=".*softwinner.*")
        
        while time.time() - start_time < duration:
            if all(v != "0.000" for v in final_data.values()):
                break
                
            win.set_focus()
            full_img = win.capture_as_image()
            
            for row_idx, row_y in enumerate(Y_STARTS):
                for i in range(12):
                    ch_id = f"CH{row_idx * 12 + i + 1:02d}"
                    if final_data[ch_id] != "0.000": continue

                    x = X_START + i * X_GAP
                    roi = full_img.crop((x, row_y, x + WIDTH, row_y + HEIGHT))
                    val_str = decimal(ocr_process(roi))
                    val_float = float(val_str)

                    # 触发条件：数值 > 0.9 且还没采满
                    if val_float > 0.900:
                        sample_pools[ch_id].append(val_float)
                        print(f"{ch_id} 正在采样: {val_float} ({len(sample_pools[ch_id])}/{max_samples})")
                        
                        # 采满 10 个数据后进行结算
                        if len(sample_pools[ch_id]) >= max_samples:
                            # 取最后 5 个数据，判断波动
                            last_five = sample_pools[ch_id][-5:]
                            # 计算极差（最大值减最小值），如果波动极小则锁定
                            if (max(last_five) - min(last_five)) < 0.005: 
                                # 取最后一段的平均值，保留三位小数
                                final_val = round(sum(last_five) / len(last_five), 3)
                                final_data[ch_id] = f"{final_val:.3f}"
                                print(f"{ch_id} 采样完成，最终值: {final_data[ch_id]}")
                            else:
                                # 如果波动还是太大，丢弃最旧的一个，继续等下一个数据点（滑动窗口）
                                sample_pools[ch_id].pop(0)
                    
                    # 如果采样中途数值突然归零，说明滤光片拔了，重置采样
                    elif val_float < 0.100 and len(sample_pools[ch_id]) > 0:
                        sample_pools[ch_id] = []
                        print(f"{ch_id} 读数归零，采样重置")

            time.sleep(interval) 

        # 导出结果
        valid_values = [float(v) for v in final_data.values() if float(v) > 0.0001]

        df_list = []
        for k, v in final_data.items():
            val_float = float(v)
            # 如果是 0.000，存为 None，Excel 会显示为空
            display_val = val_float if val_float > 0.0001 else None
            df_list.append({"通道号": k, "测量数值": display_val})
        
        df = pd.DataFrame(df_list)

        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            ws = writer.sheets['Sheet1']

            # 在 Python 中计算统计指标
            if valid_values:
                max_v = max(valid_values)
                min_v = min(valid_values)
                avg_v = sum(valid_values) / len(valid_values)
                # 计算标准差 (ddof=0 对应 Excel 的 STDEV.P)
                std_v = pd.Series(valid_values).std(ddof=0)
                cv_v = std_v / avg_v if avg_v != 0 else 0
            else:
                max_v = min_v = avg_v = cv_v = 0

            # 写入统计行（空一格后写入数值，非公式）
            stats_start_row = len(df) + 3
            stats_data = [
                ("最大值", max_v),
                ("最小值", min_v),
                ("平均值", avg_v),
                ("CV值", cv_v)
            ]

            for idx, (name, value) in enumerate(stats_data):
                curr_row = stats_start_row + idx
                ws.cell(row=curr_row, column=1, value=name)
                ws.cell(row=curr_row, column=2, value=value)
                
                # 设置样式
                ws.cell(row=curr_row, column=1).font = Font(name='微软雅黑', size=12, bold=True)
                
                if name == "CV值":
                    ws.cell(row=curr_row, column=2).number_format = '0.00%'
                else:
                    ws.cell(row=curr_row, column=2).number_format = '0.000'

            # 设置全表格式
            for r in range(2, len(df) + 2):
                ws.cell(row=r, column=2).number_format = '0.000'
                
            for row in ws.iter_rows():
                for cell in row:
                    # 获取当前的加粗状态，确保“最大值”等标题的加粗不被覆盖
                    is_bold = cell.font.bold if cell.font else False
                    
                    cell.font = Font(
                        name='微软雅黑', 
                        size=12, 
                        bold=is_bold  # 保留原有的加粗设置
                    )
        
        print(f"\n数据已保存至: {os.path.abspath(file_name)}")

    except Exception as e:
        print(f"报错: {e}")

if __name__ == "__main__":
    run()