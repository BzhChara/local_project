import time
import pandas as pd
import numpy as np
from datetime import datetime
from pywinauto import Application
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# 配置区
EXE_NAME = "Weigh.exe"
OUTPUT_FILE = "24通道原始值记录.xlsx"
INTERVAL = 3

def get_val(ctrl):
    try:
        return (ctrl.get_value() or ctrl.window_text() or "").strip()
    except:
        return ""

def apply_style(file_path):
    # 设置 Excel 格式：微软雅黑, 12号, 居中, 首行加粗
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        # 1. 定义基本样式
        normal_font = Font(name='微软雅黑', size=12, bold=False)
        header_font = Font(name='微软雅黑', size=12, bold=True) # 表头专用加粗
        my_align = Alignment(horizontal='center', vertical='center')

        # 2. 遍历单元格应用样式
        for row_idx, row in enumerate(ws.iter_rows(min_row=1), start=1):
            for cell in row:
                # 如果是第一行，使用加粗字体
                if row_idx == 1:
                    cell.font = header_font
                else:
                    cell.font = normal_font
                
                cell.alignment = my_align
        
        wb.save(file_path)
    except Exception as e:
        print(f"设置样式失败: {e}")

def main():
    try:
        print("正在连接 Weigh.exe...")
        app = Application(backend="uia").connect(path=EXE_NAME)
        dlg = app.top_window()

        time_history = [] 
        data_history = [] 

        while True:
            current_time = datetime.now().strftime("%H:%M:%S")
            
            # 1. 抓取与排序 (核心逻辑)
            all_controls = dlg.descendants(control_type="Edit") + dlg.descendants(control_type="ComboBox")
            min_left = min([b.rectangle().left for b in all_controls])
            data_controls = [b for b in all_controls if b.rectangle().left > min_left + 150]
            data_controls.sort(key=lambda b: (round(b.rectangle().top / 5), b.rectangle().left))

            # 2. 提取 24 个小数
            current_row = []
            for ctrl in data_controls:
                val = get_val(ctrl)
                if "." in val:
                    try:
                        current_row.append(float(val))
                    except: continue
            
            current_row = current_row[:24]
            if len(current_row) < 24:
                current_row += [0.0] * (24 - len(current_row))

            # 3. 存入历史记录
            time_history.append(current_time)
            data_history.append(current_row)
            data_np = np.array(data_history)

            # 4. 计算统计值 (对齐 Excel STDEV.P)
            means = np.mean(data_np, axis=0)
            stds = np.std(data_np, axis=0, ddof=0)
            cvs = [(s / m * 100) if m != 0 else 0 for m, s in zip(means, stds)]

            # 5. 构造展示 DataFrame
            df_main = pd.DataFrame(data_history, columns=[f'CH{i+1}' for i in range(24)])
            df_main.insert(0, '时间', time_history)
            
            # 构造一个与数据列数一致的空行 DataFrame
            empty_row = pd.DataFrame([[None] * len(df_main.columns)], columns=df_main.columns)

            # 统计行
            mean_row = ["平均值"] + [round(m, 6) for m in means]
            cv_row = ["CV值"] + [f"{round(c, 4)}%" for c in cvs]
            
            stat_df = pd.DataFrame([mean_row, cv_row], columns=df_main.columns)
            # 按顺序拼接：原始数据 + 空行 + 统计结果
            final_df = pd.concat([df_main, empty_row, stat_df], ignore_index=True)

            # 6. 保存
            try:
                final_df.to_excel(OUTPUT_FILE, index=False)
                apply_style(OUTPUT_FILE) # 调用加粗函数
                print(f"[{current_time}] 已保存。CH1 CV: {cv_row[1]} ...")
            except PermissionError:
                print(f"[{current_time}] Excel 被占用，请关闭文件！")

            time.sleep(INTERVAL)

    except Exception as e:
        print(f"出错: {e}")

if __name__ == "__main__":
    main()